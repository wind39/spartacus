import collections
import openpyxl
import bs4

import Spartacus.Database
import Spartacus.Report

try:

    v_reportName = "List of Employees"
    v_sheetName = "Employees"
    v_reportFile = "employees.xlsx"

    v_database = Spartacus.Database.SQLite("samples/employees.db")
    v_table = v_database.Query(
        """
        select emp_no,
               birth_date,
               first_name,
               last_name,
               gender,
               hire_date,
               abs(random() % 5) as num_dependents,
               abs(random() / 18446744073709551616 * 10000) as salary,
               '=#column_salary##row# / (#column_num_dependents##row# + 1)' as formula_salary_per_capita
        from employees
    """
    )
    if len(v_table.Rows) == 0:
        exit(0)

    v_workBook = openpyxl.Workbook()
    v_workBook.remove(v_workBook.active)

    v_workSheet = v_workBook.create_sheet(v_sheetName)
    v_workSheet.sheet_view.showGridLines = False

    # v_workSheet.add_image(
    #    openpyxl.drawing.image.Image('/path/to/image.png'),
    #    'A1'
    # )

    v_boldFont = openpyxl.styles.Font(bold=True)

    v_centerAlignment = openpyxl.styles.Alignment(
        horizontal="center", vertical="center", wrapText=True
    )

    v_workSheet.merge_cells("A2:F2")
    v_workSheet["A2"] = bs4.BeautifulSoup(v_reportName, "lxml").get_text()
    v_workSheet["A2"].font = v_boldFont
    v_workSheet["A2"].alignment = v_centerAlignment

    v_headerFont = openpyxl.styles.Font(bold=True)

    v_headerFill = openpyxl.styles.PatternFill("solid", fgColor="DBE5F1")

    v_headerAlignment = openpyxl.styles.Alignment(
        horizontal="center", vertical="center", wrapText=True
    )

    v_numberAlignment = openpyxl.styles.Alignment(
        horizontal="right", vertical="center", wrapText=True
    )

    v_textAlignment = openpyxl.styles.Alignment(
        horizontal="left", vertical="center", wrapText=True
    )

    v_dateAlignment = openpyxl.styles.Alignment(
        horizontal="center", vertical="center", wrapText=True
    )

    v_headerDict = collections.OrderedDict()

    v_headerDict["emp_no"] = Spartacus.Report.Field(
        p_name="Employee Number",
        p_width=15,
        p_comment=None,
        p_border=None,
        p_font=v_headerFont,
        p_fill=v_headerFill,
        p_alignment=v_headerAlignment,
        p_data=Spartacus.Report.Data(
            p_type="int",
            p_border=None,
            p_font=None,
            p_fill=None,
            p_alignment=v_numberAlignment,
        ),
        p_summaryList=[],
    )

    v_headerDict["first_name"] = Spartacus.Report.Field(
        p_name="First Name",
        p_width=30,
        p_comment=None,
        p_border=None,
        p_font=v_headerFont,
        p_fill=v_headerFill,
        p_alignment=v_headerAlignment,
        p_data=Spartacus.Report.Data(
            p_type="str",
            p_border=None,
            p_font=None,
            p_fill=None,
            p_alignment=v_textAlignment,
        ),
        p_summaryList=[],
    )

    v_headerDict["last_name"] = Spartacus.Report.Field(
        p_name="Last Name",
        p_width=30,
        p_comment=None,
        p_border=None,
        p_font=v_headerFont,
        p_fill=v_headerFill,
        p_alignment=v_headerAlignment,
        p_data=Spartacus.Report.Data(
            p_type="str",
            p_border=None,
            p_font=None,
            p_fill=None,
            p_alignment=v_textAlignment,
        ),
        p_summaryList=[],
    )

    v_headerDict["gender"] = Spartacus.Report.Field(
        p_name="Gender",
        p_width=8,
        p_comment=None,
        p_border=None,
        p_font=v_headerFont,
        p_fill=v_headerFill,
        p_alignment=v_headerAlignment,
        p_data=Spartacus.Report.Data(
            p_type="str",
            p_border=None,
            p_font=None,
            p_fill=None,
            p_alignment=v_textAlignment,
        ),
        p_summaryList=[],
    )

    v_headerDict["birth_date"] = Spartacus.Report.Field(
        p_name="Birth Date",
        p_width=15,
        p_comment=None,
        p_border=None,
        p_font=v_headerFont,
        p_fill=v_headerFill,
        p_alignment=v_headerAlignment,
        p_data=Spartacus.Report.Data(
            p_type="date",
            p_border=None,
            p_font=None,
            p_fill=None,
            p_alignment=v_dateAlignment,
        ),
        p_summaryList=[],
    )

    v_headerDict["hire_date"] = Spartacus.Report.Field(
        p_name="Hire Date",
        p_width=15,
        p_comment=None,
        p_border=None,
        p_font=v_headerFont,
        p_fill=v_headerFill,
        p_alignment=v_headerAlignment,
        p_data=Spartacus.Report.Data(
            p_type="date",
            p_border=None,
            p_font=None,
            p_fill=None,
            p_alignment=v_dateAlignment,
        ),
        p_summaryList=[],
    )

    v_headerDict["num_dependents"] = Spartacus.Report.Field(
        p_name="Num. Dependents",
        p_width=20,
        p_comment=None,
        p_border=None,
        p_font=v_headerFont,
        p_fill=v_headerFill,
        p_alignment=v_headerAlignment,
        p_data=Spartacus.Report.Data(
            p_type="int",
            p_border=None,
            p_font=None,
            p_fill=None,
            p_alignment=v_numberAlignment,
        ),
        p_summaryList=[
            Spartacus.Report.Summary(
                p_type="int",
                p_border=None,
                p_font=None,
                p_fill=None,
                p_function="=SUM(#column##start_row#:#column##end_row#)",
                p_index=-2,
            ),
            Spartacus.Report.Summary(
                p_type="int",
                p_border=None,
                p_font=None,
                p_fill=None,
                p_function="=SUBTOTAL(9, #column##start_row#:#column##end_row#)",
                p_index=-1,
            ),
        ],
    )

    v_headerDict["salary"] = Spartacus.Report.Field(
        p_name="Salary",
        p_width=20,
        p_comment=None,
        p_border=None,
        p_font=v_headerFont,
        p_fill=v_headerFill,
        p_alignment=v_headerAlignment,
        p_data=Spartacus.Report.Data(
            p_type="float",
            p_border=None,
            p_font=None,
            p_fill=None,
            p_alignment=v_numberAlignment,
        ),
        p_summaryList=[
            Spartacus.Report.Summary(
                p_type="float",
                p_border=None,
                p_font=None,
                p_fill=None,
                p_function="=SUM(#column##start_row#:#column##end_row#)",
                p_index=-2,
            ),
            Spartacus.Report.Summary(
                p_type="float",
                p_border=None,
                p_font=None,
                p_fill=None,
                p_function="=SUBTOTAL(9, #column##start_row#:#column##end_row#)",
                p_index=-1,
            ),
        ],
    )

    v_headerDict["formula_salary_per_capita"] = Spartacus.Report.Field(
        p_name="Salary Per Capita",
        p_width=30,
        p_comment=None,
        p_border=None,
        p_font=v_headerFont,
        p_fill=v_headerFill,
        p_alignment=v_headerAlignment,
        p_data=Spartacus.Report.Data(
            p_type="float_formula",
            p_border=None,
            p_font=None,
            p_fill=None,
            p_alignment=v_numberAlignment,
        ),
        p_summaryList=[],
    )

    v_startPerc = 0.0
    v_endPerc = 100.0
    v_inc = 0.0
    v_totalRows = len(v_table.Rows)
    try:
        v_inc = (v_endPerc - v_startPerc) / v_totalRows
    except ZeroDivisionError:
        v_inc = v_endPerc - v_startPerc
        pass
    v_progress = float(v_startPerc - v_inc)

    for v_line in Spartacus.Report.AddTable(
        p_workSheet=v_workSheet,
        p_headerDict=v_headerDict,
        p_startColumn=1,
        p_startRow=6,
        p_headerHeight=40,
        p_data=v_table,
        p_mainTable=True,
    ):
        if v_line % 1000 == 0:
            v_progress += v_inc * 1000
            print(
                "{0}: Rendering line {1} of {2} - {3}%".format(
                    v_reportName, v_line, v_totalRows, v_progress
                )
            )

    v_workBook.save(v_reportFile)

except Spartacus.Database.Exception as exc:
    print(str(exc))
except Spartacus.Report.Exception as exc:
    print(str(exc))
except Exception as exc:
    print(str(exc))
