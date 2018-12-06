'''
The MIT License (MIT)

Copyright (c) 2014-2018 William Ivanski
Copyright (c) 2018 Israel Barth Rubio

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
'''

import collections
import openpyxl
import bs4
import re

import Spartacus
import Spartacus.Database


class Exception(Exception):
    """Class used to fire custom exceptions.

        Examples:
            raise Spartacus.Report.Exception('Problem occurred while doing something.')
    """
    pass


class Range:
    """Represents a range in a worksheet.

        Attributes:
            startRow (int): the first row of the range. Defaults to 1.
                Notes:
                    Must be a positive integer.
            startColumn (int): the first column of the range. Defaults to 1.
                Notes:
                    Must be a positive integer.
            endRow (int): the last row of the range. Defaults to 1.
                Notes:
                    Must be a positive integer.
            endColumn (int): the last column of the range. Defaults to 1.
                Notes:
                    Must be a positive integer.
    """

    def __init__(self, p_startRow, p_startColumn, p_endRow, p_endColumn):
        """Create a new Spartacus.Reports.Range instance.

            Args:
                p_startRow (int): the first row of the range. Defaults to 1.
                    Notes:
                        Must be a positive integer.
                p_startColumn (int): the first column of the range. Defaults to 1.
                    Notes:
                        Must be a positive integer.
                p_endRow (int): the last row of the range. Defaults to 1.
                    Notes:
                        Must be a positive integer.
                p_endColumn (int): the last column of the range. Defaults to 1.
                    Notes:
                        Must be a positive integer.

            Raises:
                Spartacus.Report.Exception: custom exceptions occurred in this script.
        """

        if not isinstance(p_startRow, int):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Reports.Range": Parameter "p_startRow" must be of type "int".')

        if not isinstance(p_startColumn, int):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Reports.Range": Parameter "p_startColumn" must be of type "int".')

        if not isinstance(p_endRow, int):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Reports.Range": Parameter "p_endRow" must be of type "int".')

        if not isinstance(p_endColumn, int):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Reports.Range": Parameter "p_endColumn" must be of type "int".')

        self.startRow = p_startRow
        self.startColumn = p_startColumn
        self.endRow = p_endRow
        self.endColumn = p_endColumn


class Data:
    """Data represents info about data formatting in a column.

        Attributes:
            type (str): the data type of each cell of this column. Defaults to 'str'.
                Must be one of: int, float, float4, percent, date, str, int_formula, float_formula, float4_formula, percent_formula, date_formula, str_formula.
            border (openpyxl.styles.borders.Border): the border to be applied data to cells of this column. Defaults to None.
                Examples:
                    openpyxl.styles.borders.Border(
                        left = openpyxl.styles.borders.Side(
                            style = 'thick'
                        ),
                        top = openpyxl.styles.borders.Side(
                            style = 'thick'
                        ),
                        bottom = openpyxl.styles.borders.Side(
                            style = 'thick'
                        )
                    )
            font (openpyxl.styles.Font): the font to be applied to data cells of this column. Defaults to None.
                Examples:
                    openpyxl.styles.Font(
                        bold = True
                    )
            fill (openpyxl.styles.PatternFill): the fill to be applied to data cells of this column. Defaults to None.
                Examples:
                    openpyxl.styles.PatternFill(
                        'solid',
                        fgColor = 'DBE5F1'
                    )
            alignment (openpyxl.styles.Alignment): the alignment to be applied to cells of this column. Defaults to None.
                Examples:
                    p_alignment = openpyxl.styles.Alignment(
                        horizontal = 'center',
                        vertical = 'center',
                        wrapText = True
                    )
    """

    def __init__(self, p_type = 'str', p_border = None, p_font = None, p_fill = None, p_alignment = None):
        """Create a new Spartacus.Reports.Data instance.

            Args:
                p_type (str): the data type of each cell of this column. Defaults to 'str'.
                    Must be one of: int, float, float4, percent, date, str, int_formula, float_formula, float4_formula, percent_formula, date_formula, str_formula.
                p_border (openpyxl.styles.borders.Border): the border to be applied to cells of this column. Defaults to None.
                    Examples:
                        p_border = openpyxl.styles.borders.Border(
                            left = openpyxl.styles.borders.Side(
                                style = 'thick'
                            ),
                            top = openpyxl.styles.borders.Side(
                                style = 'thick'
                            ),
                            bottom = openpyxl.styles.borders.Side(
                                style = 'thick'
                            )
                        )
                p_font (openpyxl.styles.Font): the font to be applied to cells of this column. Defaults to None.
                    Examples:
                        p_font = openpyxl.styles.Font(
                            bold = True
                        )
                p_fill (openpyxl.styles.PatternFill): the fill to be applied to cells of this column. Defaults to None.
                    Examples:
                        p_fill = openpyxl.styles.PatternFill(
                            'solid',
                            fgColor = 'DBE5F1'
                        )
                p_alignment (openpyxl.styles.Alignment): the alignment to be applied to cells of this column. Defaults to None.
                    Examples:
                        p_alignment = openpyxl.styles.Alignment(
                            horizontal = 'center',
                            vertical = 'center',
                            wrapText = True
                        )

            Raises:
                Spartacus.Report.Exception: custom exceptions occurred in this script.
        """

        if not isinstance(p_type, str):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Reports.Data": Parameter "p_type" must be of type "str".')

        if not p_type in ['int', 'float', 'float4', 'percent', 'date', 'str', 'int_formula', 'float_formula', 'float4_formula', 'percent_formula', 'date_formula', 'str_formula']:
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Reports.Data": Parameter "p_type" must be one of: "int", "float", "float4", "percent", "date", "str", "int_formula", "float_formula", "float4_formula", "percent_formula", "date_formula", "str_formula".')

        if p_border is not None and not isinstance(p_border, openpyxl.styles.borders.Border):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Reports.Data": Parameter "p_border" must be None or of type "openpyxl.styles.borders.Border".')

        if p_font is not None and not isinstance(p_font, openpyxl.styles.Font):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Reports.Data": Parameter "p_font" must be None or of type "openpyxl.styles.Font".')

        if p_fill is not None and not isinstance(p_fill, openpyxl.styles.PatternFill):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Reports.Data": Parameter "p_fill" must be None or of type "openpyxl.styles.PatternFill".')

        if p_alignment is not None and not isinstance(p_alignment, openpyxl.styles.Alignment):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Reports.Data": Parameter "p_alignment" must be of type "openpyxl.styles.Alignment".')

        self.type = p_type
        self.border = p_border
        self.font = p_font
        self.fill = p_fill
        self.alignment = p_alignment


class Summary:
    """Summary represents info about one summary to be applied to a column, including type and formatting.

        Attributes:
            type (str): the data type of the column summary. Must be one of: int, float, float4, percent. Defaults to 'float'.
            border (openpyxl.styles.borders.Border): the border to be applied to column summary. Defaults to None.
                Examples:
                    openpyxl.styles.borders.Border(
                        left = openpyxl.styles.borders.Side(
                            style = 'thick'
                        ),
                        top = openpyxl.styles.borders.Side(
                            style = 'thick'
                        ),
                        bottom = openpyxl.styles.borders.Side(
                            style = 'thick'
                        )
                    )
            font (openpyxl.styles.Font): the font to be applied to column summary. Defaults to None.
                Examples:
                    openpyxl.styles.Font(
                        bold = True
                    )
            fill (openpyxl.styles.PatternFill): the fill to be applied to column summary. Defaults to None.
                Examples:
                    openpyxl.styles.PatternFill(
                        'solid',
                        fgColor = 'DBE5F1'
                    )
            function (str): it is the excel function of the summary. Defaults to ''.
                Notes: Some wildcards may be used, and will be replaced by the corresponding values when the table is built:
                    #column#: will be replaced by the column corresponding to where this summary was placed in excel.
                    #start_row#: will be replaced by the first data line number of the table this summary belongs to.
                    #end_row#: will be replaced by the last data line number of the table this summary belongs to.
                    #column_columname#: will be replaced by the letter of the column.
                Examples:
                    '=SUM(#column##start_row#:#column##end_row#)'. This example use all available wildcards.
            index (int): the line relative to the table where this summary should be placed. Defaults to -1.
                Notes:
                    Must be a positive or a negative number.
                        If index is negative, summary will be placed index lines before table header.
                        If index is positive, summary will be placed index lines after table last data line.
    """

    def __init__(self, p_type = 'float', p_border = None, p_font = None, p_fill = None, p_function = '', p_index = -1):
        """Create a new Spartacus.Reports.Summary instance.

            Args:
                p_type (str): the data type of the column summary. Must be one of: int, float, float4, percent. Defaults to 'float'.
                p_border (openpyxl.styles.borders.Border): the border to be applied to column summary. Defaults to None.
                    Examples:
                        p_border = openpyxl.styles.borders.Border(
                            left = openpyxl.styles.borders.Side(
                                style = 'thick'
                            ),
                            top = openpyxl.styles.borders.Side(
                                style = 'thick'
                            ),
                            bottom = openpyxl.styles.borders.Side(
                                style = 'thick'
                            )
                        )
                p_font (openpyxl.styles.Font): the font to be applied to column summary. Defaults to None.
                    Examples:
                        p_font = openpyxl.styles.Font(
                            bold = True
                        )
                p_fill (openpyxl.styles.PatternFill): the fill to be applied to column summary. Defaults to None.
                    Examples:
                        p_fill = openpyxl.styles.PatternFill(
                            'solid',
                            fgColor = 'DBE5F1'
                        )
                p_function (str): it is the excel function of the summary. Defaults to ''.
                    Notes: Some wildcards may be used, and will be replaced by the corresponding values when the table is built:
                        #column#: will be replaced by the column corresponding to where this summary was placed in excel.
                        #start_row#: will be replaced by the first data line number of the table this summary belongs to.
                        #end_row#: will be replaced by the last data line number of the table this summary belongs to.
                        #column_columname#: will be replaced by the letter of the column.
                    Examples:
                        p_function = '=SUM(#column##start_row#:#column##end_row#)'. This example use all available wildcards.
                p_index (int): the line relative to the table where this summary should be placed. Defaults to -1.
                    Notes:
                        Must be a positive or a negative number.
                            If index is negative, summary will be placed index lines before table header.
                            If index is positive, summary will be placed index lines after table last data line.

            Raises:
                Spartacus.Reports.Exception: custom exceptions occurred in this script.
        """

        if not isinstance(p_type, str):
            raise Spartacus.Reports.Exception('Error during instantiation of class "Spartacus.Reports.Summary": Parameter "p_type" must be of type "str".')

        if not p_type in ['int', 'float', 'float4', 'percent']:
            raise Spartacus.Reports.Exception('Error during instantiation of class "Spartacus.Reports.Summary": Parameter "p_type" must be one of: "int", "float", "float4", "percent".')

        if p_border is not None and not isinstance(p_border, openpyxl.styles.borders.Border):
            raise Spartacus.Reports.Exception('Error during instantiation of class "Spartacus.Reports.Summary": Parameter "p_border" must be None or of type "openpyxl.styles.borders.Border".')

        if p_font is not None and not isinstance(p_font, openpyxl.styles.Font):
            raise Spartacus.Reports.Exception('Error during instantiation of class "Spartacus.Reports.Summary": Parameter "p_font" must be None or of type "openpyxl.styles.Font".')

        if p_fill is not None and not isinstance(p_fill, openpyxl.styles.PatternFill):
            raise Spartacus.Reports.Exception('Error during instantiation of class "Spartacus.Reports.Summary": Parameter "p_fill" must be None or of type "openpyxl.styles.PatternFill".')

        if not isinstance(p_function, str):
            raise Spartacus.Reports.Exception('Error during instantiation of class "Spartacus.Reports.Summary": Parameter "p_function" must be of type "str".')

        if not isinstance(p_index, int):
            raise Spartacus.Reports.Exception('Error during instantiation of class "Spartacus.Reports.Summary": Parameter "p_index" must be of type "int".')

        if p_index == 0:
            raise Spartacus.Reports.Exception('Error during instantiation of class "Spartacus.Reports.Summary": Parameter "p_index" must be positive or negative.')

        self.type = p_type
        self.border = p_border
        self.font = p_font
        self.fill = p_fill
        self.function = p_function
        self.index = p_index


class Field:
    """Field represents info about one table column.

        Attributes:
            name (str): the column header name. Defaults to ''.
            width (float): the column width in pt. Defaults to 18.
                Notes: Must be a non-negative number.
            comment (openpyxl.comments.Comment): a comment that should be placed in the column header. Defaults to None.
                Examples:
                    openpyxl.comments.Comment(
                        'Comment Text',
                        'Comment Author'
                    )
            border (openpyxl.styles.borders.Border): the border to be applied to column header. Defaults to None.
                Examples:
                    openpyxl.styles.borders.Border(
                        left = openpyxl.styles.borders.Side(
                            style = 'thick'
                        ),
                        top = openpyxl.styles.borders.Side(
                            style = 'thick'
                        ),
                        bottom = openpyxl.styles.borders.Side(
                            style = 'thick'
                        )
                    )
            font (openpyxl.styles.Font): the font to be applied to column header. Defaults to None.
                Examples:
                    openpyxl.styles.Font(
                        bold = True
                    )
            fill (openpyxl.styles.PatternFill): the fill to be applied to column header. Defaults to None.
                Examples:
                    openpyxl.styles.PatternFill(
                        'solid',
                        fgColor = 'DBE5F1'
                    )
            alignment (openpyxl.styles.Alignment): the alignment to be applied to column header. Defaults to None.
                Examples:
                    openpyxl.styles.Alignment(
                        horizontal = 'center',
                        vertical = 'center',
                        wrapText = True
                    )
            data (Spartacus.Report.Data): represents formatting for column data cells. Defaults to Spartacus.Report.Data default.
                Examples:
                    Data(
                        p_type = 'int',
                        p_border = None,
                        p_font = None,
                        p_fill = None
                    )
            summaryList (list): a list where each object is an instance of Spartacus.Report.Summary. Defaults to [].
                Examples:
                    [
                        Summary(
                            p_type = 'float',
                            p_border = None,
                            p_font = None,
                            p_fill = None,
                            p_function = '=SUM(#column##start_row#:#column##end_row#)',
                            p_index = -2
                        ),
                        Summary(
                            p_type = 'float',
                            p_border = None,
                            p_font = None,
                            p_fill = None,
                            p_function = '=SUBTOTAL(9, #column##start_row#:#column##end_row#)',
                            p_index = -1
                        )
                    ]
            hidden (bool): if this column should be hidden or not. Defaults to False.
    """

    def __init__(self, p_name = '', p_width = 18, p_comment = None, p_border = None, p_font = None, p_fill = None, p_alignment = None, p_data = Data(), p_summaryList = [], p_hidden = False):
        """Create a new Spartacus.Report.Field instance.

            Args:
                p_name (str): the column header name. Defaults to ''.
                p_width (float): the column width in pt. Defaults to 18.
                    Notes: Must be a non-negative number or None.
                p_comment (openpyxl.comments.Comment): a comment that should be placed in the column header. Defaults to None.
                    Examples:
                        p_comment = openpyxl.comments.Comment(
                            'Comment Text',
                            'Comment Author'
                        )
                p_border (openpyxl.styles.borders.Border): the border to be applied to column header. Defaults to None.
                    Examples:
                        p_border = openpyxl.styles.borders.Border(
                            left = openpyxl.styles.borders.Side(
                                style = 'thick'
                            ),
                            top = openpyxl.styles.borders.Side(
                                style = 'thick'
                            ),
                            bottom = openpyxl.styles.borders.Side(
                                style = 'thick'
                            )
                        )
                p_font (openpyxl.styles.Font): the font to be applied to column header. Defaults to None.
                    Examples:
                        p_font = openpyxl.styles.Font(
                            bold = True
                        )
                p_fill (openpyxl.styles.PatternFill): the fill to be applied to column header. Defaults to None.
                    Examples:
                        p_fill = openpyxl.styles.PatternFill(
                            'solid',
                            fgColor = 'DBE5F1'
                        )
                p_alignment (openpyxl.styles.Alignment): the alignment to be applied to column header. Defaults to None.
                    Examples:
                        p_alignment = openpyxl.styles.Alignment(
                            horizontal = 'center',
                            vertical = 'center',
                            wrapText = True
                        )
                p_data (Spartacus.Report.Data): represents formatting for column data cells. Defaults to Spartacus.Report.Data default.
                    Examples:
                        p_data = Data(
                            p_type = 'int',
                            p_border = None,
                            p_font = None,
                            p_fill = None
                        )
                p_summaryList (list): a list where each object is an instance of Spartacus.Report.Summary. Defaults to [].
                    Examples:
                        p_summaryList = [
                            Summary(
                                p_type = 'float',
                                p_border = None,
                                p_font = None,
                                p_fill = None,
                                p_function = '=SUM(#column##start_row#:#column##end_row#)',
                                p_index = -2
                            ),
                            Summary(
                                p_type = 'float',
                                p_border = None,
                                p_font = None,
                                p_fill = None,
                                p_function = '=SUBTOTAL(9, #column##start_row#:#column##end_row#)',
                                p_index = -1
                            )
                        ]
                p_hidden (bool): if this column should be hidden or not. Defaults to False.

            Raises:
                Spartacus.Report.Exception: custom exceptions occurred in this script.
        """

        if not isinstance(p_name, str):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Report.Field": Parameter "p_name" must be of type "str".')

        if p_width is not None and not isinstance(p_width, int) and not isinstance(p_width, float):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Report.Field": Parameter "p_width" must be of type "int" or "float" or None.')

        if p_width is not None and p_width < 0:
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Report.Field": Parameter "p_width" must be a non-negative number.')

        if p_comment is not None and not isinstance(p_comment, openpyxl.comments.Comment):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Report.Field": Parameter "p_comment" must be of type "openpyxl.comments.Comment".')

        if p_border is not None and not isinstance(p_border, openpyxl.styles.borders.Border):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Report.Field": Parameter "p_border" must be of type "openpyxl.styles.borders.Border".')

        if p_font is not None and not isinstance(p_font, openpyxl.styles.Font):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Report.Field": Parameter "p_font" must be of type "openpyxl.styles.Font".')

        if p_fill is not None and not isinstance(p_fill, openpyxl.styles.PatternFill):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Report.Field": Parameter "p_fill" must be of type "openpyxl.styles.PatternFill".')

        if p_alignment is not None and not isinstance(p_alignment, openpyxl.styles.Alignment):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Report.Field": Parameter "p_alignment" must be of type "openpyxl.styles.Alignment".')

        if p_data is not None and not isinstance(p_data, Data):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Report.Field": Parameter "p_data" must be of type "Spartacus.Report.Data".')

        if not isinstance(p_summaryList, list):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Report.Field": Parameter "p_summaryList" must be of type "list".')

        if not isinstance(p_hidden, bool):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Report.Field": Parameter "p_hidden" must be of type "bool".')

        self.name = p_name
        self.width = p_width
        self.comment = p_comment
        self.border = p_border
        self.font = p_font
        self.fill = p_fill
        self.alignment = p_alignment
        self.data = p_data
        self.summaryList = p_summaryList
        self.hidden = p_hidden


class ConditionalFormatting:
    """ConditionalFormatting represents info about a table conditional formatting.

        Attributes:
            formula (str): the formula that decides when to format or not. Defaults to ''.
                Notes:
                    A wildcard may be used, and will be replaced by the corresponding value when the table is built:
                        #row#: the current row.
                        #column_columname#: will be replaced by the letter of the column.
                Examples:
                    '$Y#row# = 2'
            differentialStyle (openpyxl.styles.differential.DifferentialStyle): the format to be applied when formula returns true. Defaults to None.
                Examples:
                    openpyxl.styles.differential.DifferentialStyle(
                        fill = openpyxl.styles.PatternFill(
                            bgColor = 'D3D3D3'
                        )
                    )
    """

    def __init__(self, p_formula = '', p_differentialStyle = None):
        """Create a new classes.ConditionalFormatting instance.

            Args:
                p_formula (str): the formula that decides when to format or not. Defaults to ''.
                    Notes:
                        A wildcard may be used, and will be replaced by the corresponding value when the table is built:
                            #row#: the current row.
                            #column_columname#: will be replaced by the letter of the column.
                    Examples:
                        p_formula = '$Y#row# = 2'
                p_differentialStyle (openpyxl.styles.differential.DifferentialStyle): the format to be applied when formula returns true. Defaults to None.
                    Examples:
                        p_differentialStyle = openpyxl.styles.differential.DifferentialStyle(
                            fill = openpyxl.styles.PatternFill(
                                bgColor = 'D3D3D3'
                            )
                        )

            Raises:
                Spartacus.Report.Exception: custom exceptions occurred in this script.
        """

        if not isinstance(p_formula, str):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Report.ConditionalFormatting": Parameter "p_formula" must be of type "str".')

        if p_differentialStyle is not None and not isinstance(p_differentialStyle, openpyxl.styles.differential.DifferentialStyle):
            raise Spartacus.Report.Exception('Error during instantiation of class "Spartacus.Report.ConditionalFormatting": Parameter "p_formula" must be of type "openpyxl.styles.differential.DifferentialStyle".')

        self.formula = p_formula
        self.differentialStyle = p_differentialStyle


def AddTable(p_workSheet = None, p_headerDict = None, p_startColumn = 1, p_startRow = 1, p_headerHeight = None, p_data = None, p_mainTable = False, p_conditionalFormatting = None, p_tableStyleInfo = None, p_withFilters = True):
    """Insert a table in a given worksheet.

        Args:
            p_workSheet (openpyxl.worksheet.worksheet.Worksheet): the worksheet where the table will be inserted. Defaults to None.
            p_headerDict (collections.OrderedDict): an ordered dict that contains table header columns.
                Notes:
                    Each entry is in the following form:
                        Key: Name of the column to be searched in p_data.Columns.
                        Value: Spartacus.Report.Field instance.
                Examples:
                    p_headerDict = collections.OrderedDict([
                        (
                            'field_one',
                            Field(
                                p_name = 'Code',
                                p_width = 15,
                                p_data = Data(
                                    p_type = 'int'
                                )
                            )
                        ),
                        (
                            'field_two',
                            Field(
                                p_name = 'Result',
                                p_width = 15,
                                p_data = Data(
                                    p_type = 'int_formula'
                                )
                            )
                        )
                    ])
            p_startColumn (int): the column number where the table should start. Defaults to 1.
                Notes:
                    Must be a positive integer.
            p_startRow (int): the row number where the table should start. Defaults to 1.
                Notes:
                    Must be a positive integer.
            p_headerHeight (float): the header row height in pt. Defaults to None.
                Notes:
                    Must be a non-negative number or None.
            p_data (Spartacus.Database.DataTable): the datatable that contains the data that will be inserted into the excel table. Defaults to None.
                Notes:
                    If the corresponding column data type in p_headerDict is some kind of formula, then below wildcards can be used:
                        #row#: the current row.
                        #column_columname#: will be replaced by the letter of the column.
                Examples:
                    p_data = Spartacus.Database.DataTable that contains:
                        Columns: ['field_one', 'field_two'].
                        Rows: [
                            [
                                'HAHAHA',
                                '=if(#column_field_one##row# = "HAHAHA", 1, 0)'
                            ],
                            [
                                'HEHEHE',
                                '=if(#column_field_one##row# = "HAHAHA", 1, 0)'
                            ]
                        ]
            p_mainTable (bool): if this table is the main table of the current worksheet. Defaults to False.
                Notes:
                    If it's the main table, then it will consider p_width, p_hidden and freeze panes in the first table row. The 3 parameters are ignored otherwise.
            p_conditionalFormatting (Spartacus.Report.ConditionalFormatting): a conditional formatting that should be applied to data rows. Defaults to None.
                Notes:
                    Will be applied to all data rows of this table.
                    A wildcard can be used and be replaced properly:
                        #row#: the current data row.
                        #column_columname#: will be replaced by the letter of the column.
                Examples:
                    p_conditionalFormatting = ConditionalFormatting(
                        p_formula = '$Y#row# = 2',
                        p_differentialStyle = openpyxl.styles.differential.DifferentialStyle(
                            fill = openpyxl.styles.PatternFill(
                                bgColor = 'D3D3D3'
                            )
                        )
                    )
            p_tableStyleInfo (openpyxl.worksheet.table.TableStyleInfo): a style to be applied to this table. Defaults to None.
                Notes:
                    Will not be applied to summaries, if any.
                Examples:
                    p_tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(
                        name = 'TableStyleMedium23',
                        showFirstColumn = True,
                        showLastColumn = True,
                        showRowStripes = True,
                        showColumnStripes = False
                    )

            p_withFilters (bool): if the table must contain auto-filters.

        Yields:
            int: Every 1000 lines inserted into the table, yields actual line number.

        Raises:
            Spartacus.Report.Exception: custom exceptions occurred in this script.
    """

    if not isinstance(p_workSheet, openpyxl.worksheet.worksheet.Worksheet):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_workSheet" must be of type "openpyxl.worksheet.worksheet.Worksheet".')

    if not isinstance(p_headerDict, collections.OrderedDict):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_headerDict" must be of type "collections.OrderedDict".')

    if not isinstance(p_startColumn, int):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_startColumn" must be of type "int".')

    if p_startColumn < 1:
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_startColumn" must be a positive integer.')

    if not isinstance(p_startRow, int):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_startRow" must be of type "int".')

    if p_startRow < 1:
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_startRow" must be a positive integer.')

    if p_headerHeight is not None and not isinstance(p_headerHeight, int) and not isinstance(p_headerHeight, float):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_headerHeight" must be None or of type "int" or "float".')

    if not isinstance(p_data, Spartacus.Database.DataTable):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_data" must be of type "Spartacus.Database.DataTable".')

    if not isinstance(p_mainTable, bool):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_mainTable" must be of type "bool".')

    if p_conditionalFormatting is not None and not isinstance(p_conditionalFormatting, ConditionalFormatting):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_conditionalFormatting" must be None or of type "Spartacus.Report.ConditionalFormatting".')

    if p_tableStyleInfo is not None and not isinstance(p_tableStyleInfo, openpyxl.worksheet.table.TableStyleInfo):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_tableStyleInfo" must be None or of type "openpyxl.worksheet.table.TableStyleInfo".')

    if p_withFilters is not None and not isinstance(p_withFilters, bool):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_withFilters" must be None or of type "bool".')

    #Format Header
    if p_headerHeight is not None:
        p_workSheet.row_dimensions[p_startRow].height = p_headerHeight

    v_headerList = list(p_headerDict.keys())

    for i in range(len(v_headerList)):
        v_header = p_headerDict[v_headerList[i]]
        v_letter = openpyxl.utils.get_column_letter(i + p_startColumn)

        v_cell = p_workSheet['{0}{1}'.format(v_letter, p_startRow)]
        v_cell.value = v_header.name

        if p_mainTable:
            p_workSheet.column_dimensions[v_letter].width = v_header.width
            p_workSheet.column_dimensions[v_letter].hidden = v_header.hidden

        if v_header.comment is not None:
            v_cell.comment = v_header.comment

        if v_header.border is not None:
            v_cell.border = v_header.border

        if v_header.font is not None:
            v_cell.font = v_header.font

        if v_header.fill is not None:
            v_cell.fill = v_header.fill

        if v_header.alignment is not None:
            v_cell.alignment = v_header.alignment

    if p_mainTable:
        p_workSheet.freeze_panes = 'A{0}'.format(p_startRow + 1)

    #used in formula fields, if it's the case
    v_pattern = re.compile(r'#column_[^\n\r#]*#')

    v_line = 0

    #Fill content
    for v_row in p_data.Rows:
        v_line += 1

        for i in range(len(v_headerList)):
            v_headerData = p_headerDict[v_headerList[i]].data
            v_letter = openpyxl.utils.get_column_letter(i + p_startColumn)

            v_cell = p_workSheet['{0}{1}'.format(v_letter, v_line + p_startRow)] #Plus p_startRow to "jump" report header lines

            if v_headerData.border is not None:
                v_cell.border = v_headerData.border

            if v_headerData.font is not None:
                v_cell.font = v_headerData.font

            if v_headerData.fill is not None:
                v_cell.fill = v_headerData.fill

            if v_headerData.alignment is not None:
                v_cell.alignment = v_headerData.alignment

            if v_headerData.type == 'int':
                try:
                    v_cell.value = int(v_row[v_headerList[i]]) or 0
                except Exception as exc:
                    v_cell.value = v_row[v_headerList[i]] or 0

                v_cell.number_format = '0'
            elif v_headerData.type == 'float':
                try:
                    v_cell.value = float(v_row[v_headerList[i]]) or 0.0
                except Exception as exc:
                    v_cell.value = v_row[v_headerList[i]] or 0.0

                v_cell.number_format = '#,##0.00'
            elif v_headerData.type == 'float4':
                try:
                    v_cell.value = float(v_row[v_headerList[i]]) or 0.0
                except Exception as exc:
                    v_cell.value = v_row[v_headerList[i]] or 0.0

                v_cell.number_format = '#,##0.0000'
            elif v_headerData.type == 'percent':
                try:
                    v_cell.value = float(v_row[v_headerList[i]]) or 0.0
                except Exception as exc:
                    v_cell.value = v_row[v_headerList[i]] or 0.0

                v_cell.number_format = '0.00%'
            elif v_headerData.type == 'date':
                v_cell.value = v_row[v_headerList[i]] or ''
                v_cell.number_format = 'DD/MM/YYYY'
            elif v_headerData.type == 'str':
                v_cell.value = v_row[v_headerList[i]] or ''
            if v_headerData.type == 'int_formula':
                v_value = v_row[v_headerList[i]].replace('#row#', str(p_startRow + v_line))
                v_match = re.search(v_pattern, v_value)

                while v_match is not None:
                    v_start = v_match.start()
                    v_end = v_match.end()
                    v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                    v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                    v_match = re.search(v_pattern, v_value)

                v_cell.value = v_value
                v_cell.number_format = '0'
            elif v_headerData.type == 'float_formula':
                v_value = v_row[v_headerList[i]].replace('#row#', str(p_startRow + v_line))
                v_match = re.search(v_pattern, v_value)

                while v_match is not None:
                    v_start = v_match.start()
                    v_end = v_match.end()
                    v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                    v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                    v_match = re.search(v_pattern, v_value)

                v_cell.value = v_value
                v_cell.number_format = '#,##0.00'
            elif v_headerData.type == 'float4_formula':
                v_value = v_row[v_headerList[i]].replace('#row#', str(p_startRow + v_line))
                v_match = re.search(v_pattern, v_value)

                while v_match is not None:
                    v_start = v_match.start()
                    v_end = v_match.end()
                    v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                    v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                    v_match = re.search(v_pattern, v_value)

                v_cell.value = v_value
                v_cell.number_format = '#,##0.0000'
            elif v_headerData.type == 'percent_formula':
                v_value = v_row[v_headerList[i]].replace('#row#', str(p_startRow + v_line))
                v_match = re.search(v_pattern, v_value)

                while v_match is not None:
                    v_start = v_match.start()
                    v_end = v_match.end()
                    v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                    v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                    v_match = re.search(v_pattern, v_value)

                v_cell.value = v_value
                v_cell.number_format = '0.00%'
            elif v_headerData.type == 'date_formula':
                v_value = v_row[v_headerList[i]].replace('#row#', str(p_startRow + v_line))
                v_match = re.search(v_pattern, v_value)

                while v_match is not None:
                    v_start = v_match.start()
                    v_end = v_match.end()
                    v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                    v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                    v_match = re.search(v_pattern, v_value)

                v_cell.value = v_value
                v_cell.number_format = 'DD/MM/YYYY'
            elif v_headerData.type == 'str_formula':
                v_value = v_row[v_headerList[i]].replace('#row#', str(p_startRow + v_line))
                v_match = re.search(v_pattern, v_value)

                while v_match is not None:
                    v_start = v_match.start()
                    v_end = v_match.end()
                    v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                    v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                    v_match = re.search(v_pattern, v_value)

                v_cell.value = v_value

        if v_line % 1000 == 0:
            yield v_line

    v_lastLine = len(p_data.Rows) + p_startRow

    #Apply conditional formatting, if any
    if p_conditionalFormatting is not None:
        v_startLetter = openpyxl.utils.get_column_letter(p_startColumn)
        v_finalLetter = openpyxl.utils.get_column_letter(len(v_headerList) + p_startColumn - 1)

        v_formula = p_conditionalFormatting.formula.replace('#row#', str(p_startRow + 1))

        v_match = re.search(v_pattern, v_formula)

        while v_match is not None:
            v_start = v_match.start()
            v_end = v_match.end()
            v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_formula[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
            v_formula = v_formula[:v_start] + v_matchColumn + v_formula[v_end:]
            v_match = re.search(v_pattern, v_formula)

        v_rule = openpyxl.formatting.rule.Rule(
            type = 'expression',
            formula = [v_formula],
            dxf = p_conditionalFormatting.differentialStyle
        )

        p_workSheet.conditional_formatting.add(
            '{0}{1}:{2}{3}'.format(v_startLetter, p_startRow + 1, v_finalLetter, v_lastLine),
            v_rule
        )

    #Build Summary
    for i in range(len(v_headerList)):
        v_headerSummaryList = p_headerDict[v_headerList[i]].summaryList

        for v_headerSummary in v_headerSummaryList:
            v_letter = openpyxl.utils.get_column_letter(i + p_startColumn)

            v_index = p_startRow - 1

            if v_headerSummary.index < 0:
                v_index = p_startRow + v_headerSummary.index
            elif v_headerSummary.index > 0:
                v_index = v_lastLine + v_headerSummary.index

            v_value = v_headerSummary.function.replace('#column#', v_letter).replace('#start_row#', str(p_startRow + 1)).replace('#end_row#', str(v_lastLine))

            v_match = re.search(v_pattern, v_value)

            while v_match is not None:
                v_start = v_match.start()
                v_end = v_match.end()
                v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                v_match = re.search(v_pattern, v_value)

            v_cell = p_workSheet['{0}{1}'.format(v_letter, v_index)]
            v_cell.value = v_value

            if v_headerSummary.border is not None:
                v_cell.border = v_headerSummary.border

            if v_headerSummary.font is not None:
                v_cell.font = v_headerSummary.font

            if v_headerSummary.fill is not None:
                v_cell.fill = v_headerSummary.fill

            if v_headerSummary.type == 'int':
                v_cell.number_format = '0'
            elif v_headerSummary.type == 'float':
                v_cell.number_format = '#,##0.00'
            elif v_headerSummary.type == 'float4':
                v_cell.number_format = '#,##0.0000'
            elif v_headerSummary.type == 'percent':
                v_cell.number_format = '0.00%'

    #Create a new table and add it to worksheet
    v_name = 'Table_{0}_{1}'.format(p_workSheet.title.replace(' ', ''), len(p_workSheet._tables) + 1) #excel doesn't accept same displayName in more than one table.
    v_name = ''.join([c for c in v_name if c.isalnum()]) #Excel doesn't accept non-alphanumeric characters.

    v_table = openpyxl.worksheet.table.Table(
        displayName = v_name,
        ref = '{0}{1}:{2}{3}'.format(
            openpyxl.utils.get_column_letter(p_startColumn),
            p_startRow,
            openpyxl.utils.get_column_letter(p_startColumn + len(v_headerList) - 1),
            v_lastLine
        )
    )

    if p_tableStyleInfo is not None:
        v_table.tableStyleInfo = p_tableStyleInfo

    if not p_withFilters:
        v_table.headerRowCount = 0

    p_workSheet.add_table(v_table)


def AddTable(p_workSheet = None, p_headerDict = None, p_startColumn = 1, p_startRow = 1, p_headerHeight = None, p_database = None, p_query = None, p_mainTable = False, p_conditionalFormatting = None, p_tableStyleInfo = None, p_withFilters = True):
    """Insert a table in a given worksheet.

        Args:
            p_workSheet (openpyxl.worksheet.worksheet.Worksheet): the worksheet where the table will be inserted. Defaults to None.
            p_headerDict (collections.OrderedDict): an ordered dict that contains table header columns.
                Notes:
                    Each entry is in the following form:
                        Key: Name of the column to be searched in p_data.Columns.
                        Value: Spartacus.Report.Field instance.
                Examples:
                    p_headerDict = collections.OrderedDict([
                        (
                            'field_one',
                            Field(
                                p_name = 'Code',
                                p_width = 15,
                                p_data = Data(
                                    p_type = 'int'
                                )
                            )
                        ),
                        (
                            'field_two',
                            Field(
                                p_name = 'Result',
                                p_width = 15,
                                p_data = Data(
                                    p_type = 'int_formula'
                                )
                            )
                        )
                    ])
            p_startColumn (int): the column number where the table should start. Defaults to 1.
                Notes:
                    Must be a positive integer.
            p_startRow (int): the row number where the table should start. Defaults to 1.
                Notes:
                    Must be a positive integer.
            p_headerHeight (float): the header row height in pt. Defaults to None.
                Notes:
                    Must be a non-negative number or None.
            p_database (Spartacus.Database.Generic): the database from where the data will be fetched. Defaults to None.
                Notes:
                    Must be already instantiated and must not be opened yet.
            p_query (str): the query to fetch the data that will be inserted into the excel table. Defaults to None.
                Notes:
                    If the corresponding column data type in p_headerDict is some kind of formula, then below wildcards can be used:
                        #row#: the current row.
                        #column_columname#: will be replaced by the letter of the column.
                Examples:
                    "SELECT field_one, '=if(#column_field_one##row# = "HAHAHA", 1, 0)' as field_two FROM sometable"
            p_mainTable (bool): if this table is the main table of the current worksheet. Defaults to False.
                Notes:
                    If it's the main table, then it will consider p_width, p_hidden and freeze panes in the first table row. The 3 parameters are ignored otherwise.
            p_conditionalFormatting (Spartacus.Report.ConditionalFormatting): a conditional formatting that should be applied to data rows. Defaults to None.
                Notes:
                    Will be applied to all data rows of this table.
                    A wildcard can be used and be replaced properly:
                        #row#: the current data row.
                        #column_columname#: will be replaced by the letter of the column.
                Examples:
                    p_conditionalFormatting = ConditionalFormatting(
                        p_formula = '$Y#row# = 2',
                        p_differentialStyle = openpyxl.styles.differential.DifferentialStyle(
                            fill = openpyxl.styles.PatternFill(
                                bgColor = 'D3D3D3'
                            )
                        )
                    )
            p_tableStyleInfo (openpyxl.worksheet.table.TableStyleInfo): a style to be applied to this table. Defaults to None.
                Notes:
                    Will not be applied to summaries, if any.
                Examples:
                    p_tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(
                        name = 'TableStyleMedium23',
                        showFirstColumn = True,
                        showLastColumn = True,
                        showRowStripes = True,
                        showColumnStripes = False
                    )

            p_withFilters (bool): if the table must contain auto-filters.

        Yields:
            int: Every 1000 lines inserted into the table, yields actual line number.

        Raises:
            Spartacus.Report.Exception: custom exceptions occurred in this script.
    """

    if not isinstance(p_workSheet, openpyxl.worksheet.worksheet.Worksheet):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_workSheet" must be of type "openpyxl.worksheet.worksheet.Worksheet".')

    if not isinstance(p_headerDict, collections.OrderedDict):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_headerDict" must be of type "collections.OrderedDict".')

    if not isinstance(p_startColumn, int):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_startColumn" must be of type "int".')

    if p_startColumn < 1:
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_startColumn" must be a positive integer.')

    if not isinstance(p_startRow, int):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_startRow" must be of type "int".')

    if p_startRow < 1:
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_startRow" must be a positive integer.')

    if p_headerHeight is not None and not isinstance(p_headerHeight, int) and not isinstance(p_headerHeight, float):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_headerHeight" must be None or of type "int" or "float".')

    if not isinstance(p_database, Spartacus.Database.Generic):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_database" must be of type "Spartacus.Database.Generic".')

    if not isinstance(p_query, str):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_query" must be of type "str".')

    if not isinstance(p_mainTable, bool):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_mainTable" must be of type "bool".')

    if p_conditionalFormatting is not None and not isinstance(p_conditionalFormatting, ConditionalFormatting):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_conditionalFormatting" must be None or of type "Spartacus.Report.ConditionalFormatting".')

    if p_tableStyleInfo is not None and not isinstance(p_tableStyleInfo, openpyxl.worksheet.table.TableStyleInfo):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_tableStyleInfo" must be None or of type "openpyxl.worksheet.table.TableStyleInfo".')

    if p_withFilters is not None and not isinstance(p_withFilters, bool):
        raise Spartacus.Report.Exception('Error during execution of method "Static.AddTable": Parameter "p_withFilters" must be None or of type "bool".')

    #Format Header
    if p_headerHeight is not None:
        p_workSheet.row_dimensions[p_startRow].height = p_headerHeight

    v_headerList = list(p_headerDict.keys())

    for i in range(len(v_headerList)):
        v_header = p_headerDict[v_headerList[i]]
        v_letter = openpyxl.utils.get_column_letter(i + p_startColumn)

        v_cell = p_workSheet['{0}{1}'.format(v_letter, p_startRow)]
        v_cell.value = v_header.name

        if p_mainTable:
            p_workSheet.column_dimensions[v_letter].width = v_header.width
            p_workSheet.column_dimensions[v_letter].hidden = v_header.hidden

        if v_header.comment is not None:
            v_cell.comment = v_header.comment

        if v_header.border is not None:
            v_cell.border = v_header.border

        if v_header.font is not None:
            v_cell.font = v_header.font

        if v_header.fill is not None:
            v_cell.fill = v_header.fill

        if v_header.alignment is not None:
            v_cell.alignment = v_header.alignment

    if p_mainTable:
        p_workSheet.freeze_panes = 'A{0}'.format(p_startRow + 1)

    #used in formula fields, if it's the case
    v_pattern = re.compile(r'#column_[^\n\r#]*#')

    p_database.Open()

    v_line = 0
    v_hasmorerecords = True
    while v_hasmorerecords:
        v_data = p_database.QueryBlock(p_query, 1000)

        if len(v_data.Rows) > 0:

            #Fill content
            for v_row in v_data.Rows:
                v_line += 1

                for i in range(len(v_headerList)):
                    v_headerData = p_headerDict[v_headerList[i]].data
                    v_letter = openpyxl.utils.get_column_letter(i + p_startColumn)

                    v_cell = p_workSheet['{0}{1}'.format(v_letter, v_line + p_startRow)] #Plus p_startRow to "jump" report header lines

                    if v_headerData.border is not None:
                        v_cell.border = v_headerData.border

                    if v_headerData.font is not None:
                        v_cell.font = v_headerData.font

                    if v_headerData.fill is not None:
                        v_cell.fill = v_headerData.fill

                    if v_headerData.alignment is not None:
                        v_cell.alignment = v_headerData.alignment

                    if v_headerData.type == 'int':
                        try:
                            v_cell.value = int(v_row[v_headerList[i]] or '0')
                        except Exception as exc:
                            v_cell.value = v_row[v_headerList[i]] or 0

                        v_cell.number_format = '0'
                    elif v_headerData.type == 'float':
                        try:
                            v_cell.value = float(v_row[v_headerList[i]] or '0.0')
                        except Exception as exc:
                            v_cell.value = v_row[v_headerList[i]] or 0.0

                        v_cell.number_format = '#,##0.00'
                    elif v_headerData.type == 'float4':
                        try:
                            v_cell.value = float(v_row[v_headerList[i]] or '0.0')
                        except Exception as exc:
                            v_cell.value = v_row[v_headerList[i]] or 0.0

                        v_cell.number_format = '#,##0.0000'
                    elif v_headerData.type == 'percent':
                        try:
                            v_cell.value = float(v_row[v_headerList[i]] or '0.0')
                        except Exception as exc:
                            v_cell.value = v_row[v_headerList[i]] or 0.0

                        v_cell.number_format = '0.00%'
                    elif v_headerData.type == 'date':
                        v_cell.value = v_row[v_headerList[i]] or ''
                        v_cell.number_format = 'DD/MM/YYYY'
                    elif v_headerData.type == 'str':
                        v_cell.value = v_row[v_headerList[i]] or ''
                    if v_headerData.type == 'int_formula':
                        v_value = v_row[v_headerList[i]].replace('#row#', str(p_startRow + v_line))
                        v_match = re.search(v_pattern, v_value)

                        while v_match is not None:
                            v_start = v_match.start()
                            v_end = v_match.end()
                            v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                            v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                            v_match = re.search(v_pattern, v_value)

                        v_cell.value = v_value
                        v_cell.number_format = '0'
                    elif v_headerData.type == 'float_formula':
                        v_value = v_row[v_headerList[i]].replace('#row#', str(p_startRow + v_line))
                        v_match = re.search(v_pattern, v_value)

                        while v_match is not None:
                            v_start = v_match.start()
                            v_end = v_match.end()
                            v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                            v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                            v_match = re.search(v_pattern, v_value)

                        v_cell.value = v_value
                        v_cell.number_format = '#,##0.00'
                    elif v_headerData.type == 'float4_formula':
                        v_value = v_row[v_headerList[i]].replace('#row#', str(p_startRow + v_line))
                        v_match = re.search(v_pattern, v_value)

                        while v_match is not None:
                            v_start = v_match.start()
                            v_end = v_match.end()
                            v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                            v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                            v_match = re.search(v_pattern, v_value)

                        v_cell.value = v_value
                        v_cell.number_format = '#,##0.0000'
                    elif v_headerData.type == 'percent_formula':
                        v_value = v_row[v_headerList[i]].replace('#row#', str(p_startRow + v_line))
                        v_match = re.search(v_pattern, v_value)

                        while v_match is not None:
                            v_start = v_match.start()
                            v_end = v_match.end()
                            v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                            v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                            v_match = re.search(v_pattern, v_value)

                        v_cell.value = v_value
                        v_cell.number_format = '0.00%'
                    elif v_headerData.type == 'date_formula':
                        v_value = v_row[v_headerList[i]].replace('#row#', str(p_startRow + v_line))
                        v_match = re.search(v_pattern, v_value)

                        while v_match is not None:
                            v_start = v_match.start()
                            v_end = v_match.end()
                            v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                            v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                            v_match = re.search(v_pattern, v_value)

                        v_cell.value = v_value
                        v_cell.number_format = 'DD/MM/YYYY'
                    elif v_headerData.type == 'str_formula':
                        v_value = v_row[v_headerList[i]].replace('#row#', str(p_startRow + v_line))
                        v_match = re.search(v_pattern, v_value)

                        while v_match is not None:
                            v_start = v_match.start()
                            v_end = v_match.end()
                            v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                            v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                            v_match = re.search(v_pattern, v_value)

                        v_cell.value = v_value

                yield v_line

            if p_database.v_start:
                v_hasmorerecords = False
            else:
                v_hasmorerecords = True
        else:
            v_hasmorerecords = False

    p_database.Close()

    v_lastLine = v_line + p_startRow

    #Apply conditional formatting, if any
    if p_conditionalFormatting is not None:
        v_startLetter = openpyxl.utils.get_column_letter(p_startColumn)
        v_finalLetter = openpyxl.utils.get_column_letter(len(v_headerList) + p_startColumn - 1)

        v_formula = p_conditionalFormatting.formula.replace('#row#', str(p_startRow + 1))

        v_match = re.search(v_pattern, v_formula)

        while v_match is not None:
            v_start = v_match.start()
            v_end = v_match.end()
            v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_formula[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
            v_formula = v_formula[:v_start] + v_matchColumn + v_formula[v_end:]
            v_match = re.search(v_pattern, v_formula)

        v_rule = openpyxl.formatting.rule.Rule(
            type = 'expression',
            formula = [v_formula],
            dxf = p_conditionalFormatting.differentialStyle
        )

        p_workSheet.conditional_formatting.add(
            '{0}{1}:{2}{3}'.format(v_startLetter, p_startRow + 1, v_finalLetter, v_lastLine),
            v_rule
        )

    #Build Summary
    for i in range(len(v_headerList)):
        v_headerSummaryList = p_headerDict[v_headerList[i]].summaryList

        for v_headerSummary in v_headerSummaryList:
            v_letter = openpyxl.utils.get_column_letter(i + p_startColumn)

            v_index = p_startRow - 1

            if v_headerSummary.index < 0:
                v_index = p_startRow + v_headerSummary.index
            elif v_headerSummary.index > 0:
                v_index = v_lastLine + v_headerSummary.index

            v_value = v_headerSummary.function.replace('#column#', v_letter).replace('#start_row#', str(p_startRow + 1)).replace('#end_row#', str(v_lastLine))

            v_match = re.search(v_pattern, v_value)

            while v_match is not None:
                v_start = v_match.start()
                v_end = v_match.end()
                v_matchColumn = openpyxl.utils.get_column_letter(p_startColumn + v_headerList.index(v_value[v_start + 8 : v_end - 1])) #Discard starting #column_ and ending # in match
                v_value = v_value[:v_start] + v_matchColumn + v_value[v_end:]
                v_match = re.search(v_pattern, v_value)

            v_cell = p_workSheet['{0}{1}'.format(v_letter, v_index)]
            v_cell.value = v_value

            if v_headerSummary.border is not None:
                v_cell.border = v_headerSummary.border

            if v_headerSummary.font is not None:
                v_cell.font = v_headerSummary.font

            if v_headerSummary.fill is not None:
                v_cell.fill = v_headerSummary.fill

            if v_headerSummary.type == 'int':
                v_cell.number_format = '0'
            elif v_headerSummary.type == 'float':
                v_cell.number_format = '#,##0.00'
            elif v_headerSummary.type == 'float4':
                v_cell.number_format = '#,##0.0000'
            elif v_headerSummary.type == 'percent':
                v_cell.number_format = '0.00%'

    #Create a new table and add it to worksheet
    v_name = 'Table_{0}_{1}'.format(p_workSheet.title.replace(' ', ''), len(p_workSheet._tables) + 1) #excel doesn't accept same displayName in more than one table.
    v_name = ''.join([c for c in v_name if c.isalnum()]) #Excel doesn't accept non-alphanumeric characters.

    v_table = openpyxl.worksheet.table.Table(
        displayName = v_name,
        ref = '{0}{1}:{2}{3}'.format(
            openpyxl.utils.get_column_letter(p_startColumn),
            p_startRow,
            openpyxl.utils.get_column_letter(p_startColumn + len(v_headerList) - 1),
            v_lastLine
        )
    )

    if p_tableStyleInfo is not None:
        v_table.tableStyleInfo = p_tableStyleInfo

    if not p_withFilters:
        v_table.headerRowCount = 0

    p_workSheet.add_table(v_table)
