"""
The MIT License (MIT)

Copyright (c) 2014-2019 William Ivanski
Copyright (c) 2018-2019 Israel Barth Rubio

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""

import pyscrypt
import pyaes
import base64
import os
import csv
from collections import OrderedDict
import hashlib
from enum import IntEnum
from cryptography.fernet import Fernet

v_supported_file_formats = ["csv"]
try:
    import openpyxl

    v_supported_file_formats.append("xlsx")
except ImportError:
    pass
try:
    import pyexcel
    import tempfile

    v_supported_file_formats.append("xls")
except ImportError:
    pass

import Spartacus
import Spartacus.Database


class Exception(Exception):
    pass


class CryptorBackend(IntEnum):
    AES = 1
    CRYPTOGRAPHY = 2


class Cryptor(object):
    def __init__(self, p_key, p_encoding="utf-8", p_backend=CryptorBackend.AES):
        try:
            self.v_encoding = p_encoding

            if p_backend == CryptorBackend.AES:
                self.v_hash = pyscrypt.hash(
                    password=p_key.encode("utf-8"),
                    salt="0123456789ABCDEF".encode("utf-8"),
                    N=1024,
                    r=1,
                    p=1,
                    dkLen=32,
                )
            elif p_backend == CryptorBackend.CRYPTOGRAPHY:
                if len(p_key) != 32:
                    raise Exception(
                        'p_key parameter must be exactaly 32 characters length while using "cryptography" backend'
                    )

                self.v_hash = base64.urlsafe_b64encode(p_key.encode("utf-8"))
            else:
                raise Exception("Unrecognized cryptor backend")

            self.v_backend = p_backend
        except Exception as exc:
            raise Spartacus.Utils.Exception(str(exc))

    def Encrypt(self, p_plaintext):
        try:
            if self.v_backend == CryptorBackend.AES:
                v_aes = pyaes.AESModeOfOperationCTR(self.v_hash)
                return base64.b64encode(v_aes.encrypt(p_plaintext)).decode(
                    self.v_encoding
                )
            elif self.v_backend == CryptorBackend.CRYPTOGRAPHY:
                v_fernet = Fernet(self.v_hash)
                return v_fernet.encrypt(p_plaintext.encode(self.v_encoding)).decode(
                    self.v_encoding
                )
        except Exception as exc:
            raise Spartacus.Utils.Exception(str(exc))

    def Decrypt(self, p_cyphertext):
        try:
            if self.v_backend == CryptorBackend.AES:
                v_aes = pyaes.AESModeOfOperationCTR(self.v_hash)
                return v_aes.decrypt(base64.b64decode(p_cyphertext)).decode(
                    self.v_encoding
                )
            elif self.v_backend == CryptorBackend.CRYPTOGRAPHY:
                v_fernet = Fernet(self.v_hash)
                return v_fernet.decrypt(p_cyphertext.encode(self.v_encoding)).decode(
                    self.v_encoding
                )
        except Exception as exc:
            raise Spartacus.Utils.Exception(str(exc))

    def Hash(self, p_text):
        try:
            return hashlib.md5(p_text.encode(self.v_encoding)).hexdigest()
        except Exception as exc:
            raise Spartacus.Utils.Exception(str(exc))


class DataFileReader(object):
    def __init__(
        self, p_filename, p_fieldnames=None, p_encoding="utf-8", p_delimiter=None
    ):
        v_tmp = p_filename.split(".")
        if len(v_tmp) > 1:
            self.v_extension = v_tmp[-1].lower()
        else:
            self.v_extension = "csv"
        if self.v_extension == "txt" or self.v_extension == "out":
            self.v_extension = "csv"
        self.v_filename = p_filename
        self.v_file = None
        self.v_header = p_fieldnames
        self.v_encoding = p_encoding
        self.v_delimiter = p_delimiter
        self.v_open = False
        self.v_object = None

    def Open(self):
        try:
            if not os.path.isfile(self.v_filename):
                raise Spartacus.Utils.Exception(
                    "File {0} does not exist or is not a file.".format(self.v_filename)
                )
            if self.v_extension == "csv":
                self.v_file = open(self.v_filename, encoding=self.v_encoding)
                v_sample = self.v_file.read(1024)
                self.v_file.seek(0)
                v_sniffer = csv.Sniffer()
                if not v_sniffer.has_header(v_sample):
                    raise Spartacus.Utils.Exception(
                        "CSV file {0} does not have a header.".format(self.v_filename)
                    )
                v_dialect = v_sniffer.sniff(v_sample)
                if self.v_delimiter is not None:
                    v_dialect.delimiter = self.v_delimiter
                self.v_object = csv.DictReader(
                    self.v_file, self.v_header, None, None, v_dialect
                )
                self.v_open = True
            elif self.v_extension == "xlsx":
                if 'xlsx' in v_supported_file_formats:
                    self.v_object = openpyxl.load_workbook(self.v_filename, read_only=True)
                    self.v_open = True
                else:
                    raise Spartacus.Utils.Exception(
                        "XLSX is not supported. Please install it with 'pip install Spartacus[xlsx]'."
                    )
            elif self.v_extension == "xls":
                if 'xls' in v_supported_file_formats:
                    v_tmp_file = tempfile.NamedTemporaryFile(suffix=".xlsx")
                    v_tmp_file.file.close()
                    pyexcel.save_book_as(
                        file_name=self.v_filename, dest_file_name=v_tmp_file.name
                    )
                    self.v_object = openpyxl.load_workbook(v_tmp_file.name, read_only=True)
                    self.v_open = True
                else:
                    raise Spartacus.Utils.Exception(
                        "XLS is not supported. Please install it with 'pip install Spartacus[xls]'."
                    )
            else:
                raise Spartacus.Utils.Exception(
                    'File extension "{0}" not supported.'.format(self.v_extension)
                )
        except Spartacus.Utils.Exception as exc:
            raise exc
        except Exception as exc:
            raise Spartacus.Utils.Exception(str(exc))

    def Read(self, p_blocksize=None, p_sheetname=None):
        try:
            if not self.v_open:
                raise Spartacus.Utils.Exception("You need to call Open() first.")
            if self.v_extension == "csv":
                v_table = Spartacus.Database.DataTable(None, p_alltypesstr=True)
                v_first = True
                x = 0
                for v_row in self.v_object:
                    if v_first:
                        if self.v_header:
                            v_table.Columns = self.v_header
                        else:
                            for k in v_row.keys():
                                v_table.Columns.append(k)
                        v_first = False
                    v_table.Rows.append(v_row)
                    x = x + 1
                    if x == p_blocksize:
                        yield v_table
                        x = 0
                        v_table.Rows = []
                self.v_file.close()
                if len(v_table.Rows) > 0:
                    yield v_table
            else:
                if p_sheetname:
                    v_worksheet = self.v_object[p_sheetname]
                    v_table = Spartacus.Database.DataTable(p_sheetname)
                else:
                    v_worksheet = self.v_object.active
                    v_table = Spartacus.Database.DataTable()
                v_first = True
                x = 0
                for v_row in v_worksheet.rows:
                    if v_first:
                        if self.v_header:
                            v_table.Columns = self.v_header
                        else:
                            v_table.Columns = [a.value for a in v_row]
                        v_first = False
                    else:
                        v_tmp = [a.value for a in v_row]
                        if len(v_tmp) < len(v_table.Columns):
                            for i in range(0, len(v_table.Columns) - len(v_tmp)):
                                v_tmp.append(None)
                        elif len(v_tmp) > len(v_table.Columns):
                            for i in range(0, len(v_tmp) - len(v_table.Columns)):
                                v_tmp.pop()
                        v_table.Rows.append(OrderedDict(zip(v_table.Columns, v_tmp)))
                        x = x + 1
                        if x == p_blocksize:
                            yield v_table
                            x = 0
                            v_table.Rows = []
                if len(v_table.Rows) > 0:
                    yield v_table
        except Spartacus.Utils.Exception as exc:
            raise exc
        except Exception as exc:
            raise Spartacus.Utils.Exception(str(exc))

    def Sheets(self):
        try:
            if self.v_extension in ["xlsx", "xls"] and self.v_object:
                return self.v_object.sheetnames
            else:
                return []
        except Spartacus.Utils.Exception as exc:
            raise exc
        except Exception as exc:
            raise Spartacus.Utils.Exception(str(exc))


class DataFileWriter(object):
    def __init__(
        self,
        p_filename,
        p_fieldnames=None,
        p_encoding="utf-8",
        p_delimiter=";",
        p_lineterminator="\n",
    ):
        v_tmp = p_filename.split(".")
        if len(v_tmp) > 1:
            self.v_extension = v_tmp[-1].lower()
        else:
            self.v_extension = "csv"
        if self.v_extension == "txt" or self.v_extension == "out":
            self.v_extension = "csv"
        self.v_filename = p_filename
        self.v_file = None
        self.v_header = p_fieldnames  # Can't be empty for CSV
        self.v_encoding = p_encoding
        self.v_delimiter = p_delimiter
        self.v_lineterminator = p_lineterminator
        self.v_currentrow = 1
        self.v_open = False
        self.v_object = None

    def Open(self):
        try:
            if self.v_extension == "csv":
                self.v_file = open(self.v_filename, "w", encoding=self.v_encoding)
                self.v_object = csv.writer(
                    self.v_file,
                    delimiter=self.v_delimiter,
                    lineterminator=self.v_lineterminator,
                )
                self.v_object.writerow(self.v_header)
                self.v_open = True
            elif self.v_extension == "xlsx":
                if 'xlsx' in v_supported_file_formats:
                    self.v_object = openpyxl.Workbook(write_only=True)
                    self.v_open = True
                else:
                    raise Spartacus.Utils.Exception(
                        "XLSX is not supported. Please install it with 'pip install Spartacus[xlsx]'."
                    )
            else:
                raise Spartacus.Utils.Exception(
                    'File extension "{0}" not supported.'.format(self.v_extension)
                )
        except Spartacus.Utils.Exception as exc:
            raise exc
        except Exception as exc:
            raise Spartacus.Utils.Exception(str(exc))

    def Write(self, p_datatable, p_sheetname=None):
        try:
            if not self.v_open:
                raise Spartacus.Utils.Exception("You need to call Open() first.")
            if self.v_extension == "csv":
                for v_row in p_datatable.Rows:
                    self.v_object.writerow(v_row)
            else:
                if self.v_currentrow == 1:
                    if p_sheetname:
                        v_worksheet = self.v_object.create_sheet(p_sheetname)
                    else:
                        v_worksheet = self.v_object.create_sheet()
                    v_worksheet.append(p_datatable.Columns)
                    self.v_currentrow = self.v_currentrow + 1
                else:
                    v_worksheet = self.v_object.active
                if p_datatable.Simple:
                    for r in range(0, len(p_datatable.Rows)):
                        v_row = []
                        for c in range(0, len(p_datatable.Columns)):
                            v_row.append(p_datatable.Rows[r][c])
                        v_worksheet.append(v_row)
                else:
                    for r in range(0, len(p_datatable.Rows)):
                        v_row = []
                        for c in p_datatable.Columns:
                            v_row.append(p_datatable.Rows[r][c])
                        v_worksheet.append(v_row)
                self.v_currentrow = self.v_currentrow + len(p_datatable.Rows)
        except Spartacus.Utils.Exception as exc:
            raise exc
        except Exception as exc:
            raise Spartacus.Utils.Exception(str(exc))

    def Flush(self):
        try:
            if not self.v_open:
                raise Spartacus.Utils.Exception("You need to call Open() first.")
            if self.v_extension == "csv":
                self.v_file.close()
            else:
                self.v_object.save(self.v_filename)
        except Spartacus.Utils.Exception as exc:
            raise exc
        except Exception as exc:
            raise Spartacus.Utils.Exception(str(exc))
